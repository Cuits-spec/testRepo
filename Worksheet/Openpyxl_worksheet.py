import openpyxl

# 创建对象- wb代表一个工作薄（一个excel文件）
wb = openpyxl.Workbook()
# 激活工作表-获取活跃的工作表，ws代表wb（一个工作薄）里的一个工作表
ws = wb.active
# 更改工作表的标题
ws.title = 'test_sheet1'
# 对工作表传入数据
ws['A1'] = '国家'
ws['B1'] = '首都'
data = {
    '中国': '北京',
    '韩国': '首尔',
    '日本': '东京',
    '泰国': '曼谷',
    '马来西亚': '吉隆坡',
    '越南': '河内',
    '朝鲜': '平壤',
    '印度': '新德里'
}
data_excel = []
# 将字典中的每对数据（键，值）以列表形式传入data_excel列表
for each in data:
    data_excel.append([each, data[each]])
print(data_excel)
# # 将data_excel列表内的内容存入工作表
for each in data_excel:
    ws.append(each)
# 保存文件并命名
wb.save('test.xlsx')

# 打开已经存在的excel文件
wb2 = openpyxl.load_workbook('test.xlsx')
# 激活
ws2 = wb2.active
# ws2.title = '哈哈'
print(wb2.sheetnames)